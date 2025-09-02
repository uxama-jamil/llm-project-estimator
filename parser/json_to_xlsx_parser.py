import json
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import os

def json_to_excel(json_data, output_file_path):
    """Convert JSON data to Excel matching the template structure"""
    print("🚀 Starting JSON to Excel conversion...")
    
    if not json_data:
        print("✗ No data provided")
        return False
    
    try:
        # Create workbook with single sheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Project Estimation"
        
        # Define styling
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        phase_font = Font(bold=True, size=11)
        subtotal_font = Font(bold=True, size=10)
        subtotal_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Yellow background
        grand_total_font = Font(bold=True, size=12)
        
        # Get project info
        project_info = json_data.get('projectInfo', {})
        phases = json_data.get('phases', [])
        
        current_row = 1
        
        # 1. PROJECT TITLE
        ws.cell(row=current_row, column=1, value="PROJECT ESTIMATE TEMPLATE")
        ws.cell(row=current_row, column=1).font = header_font
        ws.cell(row=current_row, column=1).fill = header_fill
        current_row += 2  # Skip row 2
        
        # 2. PROJECT INFO SECTION
        project_fields = [
            ("Project Name:", project_info.get('projectName', '')),
            ("Client:", project_info.get('client', '')),
            ("Date:", project_info.get('date', '')),
            ("Version:", project_info.get('version', '1.0')),
            ("Prepared By:", project_info.get('preparedBy', '')),
            ("Assumptions:", project_info.get('assumptions', ''))
        ]
        
        for field, value in project_fields:
            ws.cell(row=current_row, column=1, value=field)
            ws.cell(row=current_row, column=2, value=value)
            current_row += 1
        
        current_row += 2  # Add spacing
        
        # 3. TABLE HEADERS
        headers = ["Phase", "Task Name", "Resource", "Min Hours", "Max Hours", "Start Date", "End Date", "Status"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
        current_row += 1
        
        # 4. PHASES AND TASKS
        for phase in phases:
            phase_name = phase.get('name', '')
            tasks = phase.get('tasks', [])
            subtotal = phase.get('subtotal', f"{phase_name.replace('PHASE', '').strip()} Subtotal")
            
            # Phase header row
            ws.cell(row=current_row, column=1, value=phase_name)
            ws.cell(row=current_row, column=1).font = phase_font
            current_row += 1
            
            # Tasks within phase
            for task in tasks:
                ws.cell(row=current_row, column=1, value="")  # Empty phase column for tasks
                ws.cell(row=current_row, column=2, value=task.get('taskName', ''))
                ws.cell(row=current_row, column=3, value=task.get('resource', ''))
                
                # Handle Min Hours and Max Hours separately
                min_hours = task.get('minHours', '')
                max_hours = task.get('maxHours', '')
                
                # Convert to string/number if needed and display
                try:
                    if min_hours:
                        if isinstance(min_hours, (int, float)):
                            ws.cell(row=current_row, column=4, value=min_hours)
                        else:
                            # Try to convert string to number
                            min_val = float(str(min_hours)) if str(min_hours).replace('.', '').replace('-', '').isdigit() else str(min_hours)
                            ws.cell(row=current_row, column=4, value=min_val)
                    else:
                        ws.cell(row=current_row, column=4, value="")
                except (ValueError, TypeError):
                    ws.cell(row=current_row, column=4, value=str(min_hours) if min_hours else "")
                
                try:
                    if max_hours:
                        if isinstance(max_hours, (int, float)):
                            ws.cell(row=current_row, column=5, value=max_hours)
                        else:
                            # Try to convert string to number
                            max_val = float(str(max_hours)) if str(max_hours).replace('.', '').replace('-', '').isdigit() else str(max_hours)
                            ws.cell(row=current_row, column=5, value=max_val)
                    else:
                        ws.cell(row=current_row, column=5, value="")
                except (ValueError, TypeError):
                    ws.cell(row=current_row, column=5, value=str(max_hours) if max_hours else "")
                
                ws.cell(row=current_row, column=6, value=task.get('startDate', ''))
                ws.cell(row=current_row, column=7, value=task.get('endDate', ''))
                ws.cell(row=current_row, column=8, value=task.get('status', 'Not Started'))
                current_row += 1
            
            # Phase subtotal row
            # Handle subtotal - it might be a string or a dictionary with hours
            if isinstance(subtotal, dict):
                subtotal_text = f"{phase_name.replace('PHASE', '').strip()} Subtotal"
                subtotal_hours = ""
                if 'minHours' in subtotal and 'maxHours' in subtotal:
                    min_h = subtotal['minHours']
                    max_h = subtotal['maxHours']
                    if min_h != max_h:
                        subtotal_hours = f"{min_h}-{max_h}h"
                    else:
                        subtotal_hours = f"{min_h}h"
                for i in range(len(headers)):
                    ws.cell(row=current_row, column=i+1).fill = subtotal_fill
                ws.cell(row=current_row, column=1, value=subtotal_text)
                ws.cell(row=current_row, column=4, value=min_h)  # Put hours in Hours column
                ws.cell(row=current_row, column=5, value=max_h)  # Put hours in Hours column
            else:
                ws.cell(row=current_row, column=1, value=subtotal)
            ws.cell(row=current_row, column=1).font = subtotal_font
            current_row += 1
            
            # Add empty row between phases
            current_row += 1
        
        # 5. GRAND TOTAL
        grand_total = json_data.get('grandTotal', 'GRAND TOTAL')
        
        # Handle grand total - it might be a string or a dictionary
        if isinstance(grand_total, dict):
            ws.cell(row=current_row, column=1, value="GRAND TOTAL")
            if 'minHours' in grand_total and 'maxHours' in grand_total:
                min_h = grand_total['minHours']
                max_h = grand_total['maxHours']
                try:
                    ws.cell(row=current_row, column=4, value=float(min_h) if isinstance(min_h, str) and min_h.replace('.', '').isdigit() else min_h)
                    ws.cell(row=current_row, column=5, value=float(max_h) if isinstance(max_h, str) and max_h.replace('.', '').isdigit() else max_h)
                except:
                    ws.cell(row=current_row, column=4, value=str(min_h))
                    ws.cell(row=current_row, column=5, value=str(max_h))
        else:
            ws.cell(row=current_row, column=1, value=grand_total)
        
        ws.cell(row=current_row, column=1).font = grand_total_font
        
        # Calculate and display summary if available (fallback if grand total doesn't have hours)
        if not isinstance(grand_total, dict):
            summary = json_data.get('summary', {})
            if summary:
                total_min = summary.get('totalMinHours', 0)
                total_max = summary.get('totalMaxHours', 0)
                if total_min:
                    ws.cell(row=current_row, column=4, value=total_min)
                if total_max:
                    ws.cell(row=current_row, column=5, value=total_max)
        
        current_row += 3  # Add some space after grand total
        
        # 6. SUMMARY SECTION
        summary = json_data.get('summary', {})
        if summary:
            ws.cell(row=current_row, column=1, value="PROJECT SUMMARY")
            ws.cell(row=current_row, column=1).font = Font(bold=True, size=12)
            current_row += 1
            
            summary_items = [
                ("Total Min Hours:", summary.get('totalMinHours', 0)),
                ("Total Max Hours:", summary.get('totalMaxHours', 0)),
                ("Estimated Duration (Weeks):", summary.get('estimatedDurationWeeks', 0)),
                ("Recommended Team Size:", summary.get('recommendedTeamSize', 0)),
                ("Total Tasks:", summary.get('totalTasks', 0)),
                ("Total Phases:", summary.get('totalPhases', 0))
            ]
            
            for item, value in summary_items:
                if value:  # Only add non-zero values
                    ws.cell(row=current_row, column=1, value=item)
                    ws.cell(row=current_row, column=2, value=value)
                    current_row += 1
            
            current_row += 2  # Add space
        
        # 7. RISKS SECTION
        risks = json_data.get('risks', [])
        if risks:
            ws.cell(row=current_row, column=1, value="RISKS & CONSIDERATIONS")
            ws.cell(row=current_row, column=1).font = Font(bold=True, size=12)
            current_row += 1
            
            # Risk headers
            risk_headers = ["Risk", "Phase", "Impact", "Mitigation"]
            for col, header in enumerate(risk_headers, 1):
                cell = ws.cell(row=current_row, column=col, value=header)
                cell.font = Font(bold=True, size=10)
                cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
            current_row += 1
            
            # Risk data
            for risk in risks:
                ws.cell(row=current_row, column=1, value=risk.get('risk', ''))
                ws.cell(row=current_row, column=2, value=risk.get('phase', ''))
                ws.cell(row=current_row, column=3, value=risk.get('impact', ''))
                ws.cell(row=current_row, column=4, value=risk.get('mitigation', ''))
                current_row += 1
            
            current_row += 2  # Add space
        
        # 8. RECOMMENDATIONS SECTION
        recommendations = json_data.get('recommendations', [])
        if recommendations:
            ws.cell(row=current_row, column=1, value="RECOMMENDATIONS")
            ws.cell(row=current_row, column=1).font = Font(bold=True, size=12)
            current_row += 1
            
            for i, recommendation in enumerate(recommendations, 1):
                ws.cell(row=current_row, column=1, value=f"{i}.")
                ws.cell(row=current_row, column=2, value=recommendation)
                current_row += 1
            
            current_row += 2  # Add space
        
        # 9. ADJUST COLUMN WIDTHS
        column_widths = {
            'A': 35,  # Phase / Risk / Recommendations
            'B': 45,  # Task Name / Description
            'C': 20,  # Resource / Impact
            'D': 20,  # Min Hours / Mitigation
            'E': 15,  # Max Hours
            'F': 15,  # Start Date
            'G': 15,  # End Date
            'H': 15   # Status
        }
        
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
        
        # 10. APPLY BORDERS TO DATA AREA
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Find the data range (from headers to grand total)
        header_row = None
        for row in range(1, ws.max_row + 1):
            if ws.cell(row=row, column=1).value == "Phase":
                header_row = row
                break
        
        # Apply borders to main estimation table (Phase to Grand Total)
        if header_row:
            grand_total_row = None
            for row in range(header_row, ws.max_row + 1):
                if ws.cell(row=row, column=1).value == "GRAND TOTAL":
                    grand_total_row = row
                    break
            
            if grand_total_row:
                for row in range(header_row, grand_total_row + 1):
                    for col in range(1, 9):  # A to H columns (8 columns total)
                        ws.cell(row=row, column=col).border = thin_border
        
        # Apply borders to risks section if it exists
        risks_header_row = None
        for row in range(1, ws.max_row + 1):
            if ws.cell(row=row, column=1).value == "Risk":
                risks_header_row = row
                break
        
        if risks_header_row:
            risks_end_row = risks_header_row
            for row in range(risks_header_row + 1, ws.max_row + 1):
                if ws.cell(row=row, column=1).value and not ws.cell(row=row, column=2).value:
                    break
                if ws.cell(row=row, column=1).value:
                    risks_end_row = row
            
            for row in range(risks_header_row, risks_end_row + 1):
                for col in range(1, 5):  # A to D columns for risks
                    ws.cell(row=row, column=col).border = thin_border
        
        # 11. MERGE PROJECT TITLE CELL
        ws.merge_cells(f'A1:H1')
        ws.cell(row=1, column=1).alignment = Alignment(horizontal='center')
        
        # 12. SAVE WORKBOOK
        wb.save(output_file_path)
        print(f"Successfully created Excel file: {output_file_path}")
        
        # Display summary
        print("\nCONVERSION SUMMARY:")
        print(f"   • Total Phases: {len(phases)}")
        total_tasks = sum(len(phase.get('tasks', [])) for phase in phases)
        print(f"   • Total Tasks: {total_tasks}")
        
        # Show summary data
        if summary:
            print(f"   • Estimated Hours: {summary.get('totalMinHours', 0)}-{summary.get('totalMaxHours', 0)}")
            print(f"   • Duration: {summary.get('estimatedDurationWeeks', 0)} weeks")
            print(f"   • Team Size: {summary.get('recommendedTeamSize', 0)} people")
        
        # Show risks and recommendations count
        if risks:
            print(f"   • Risks Identified: {len(risks)}")
        if recommendations:
            print(f"   • Recommendations: {len(recommendations)}")
        
        return True
        
    except Exception as e:
        print(f"✗ Error creating Excel file: {e}")
        import traceback
        traceback.print_exc()
        return False

# Helper function for testing
def test_json_to_excel():
    """Test function with sample data"""
    sample_data = {
        "projectInfo": {
            "title": "Sample Mobile App Project",
            "projectName": "TaskMaster Mobile App",
            "client": "TechCorp Inc.",
            "date": "2025-01-15",
            "version": "1.0",
            "preparedBy": "John Doe",
            "assumptions": "Client will provide all required APIs and design assets"
        },
        "phases": [
            {
                "name": "DESIGN & ARCHITECTURE PHASE",
                "tasks": [
                    {
                        "taskName": "Application Architecture",
                        "resource": "Senior Architect",
                        "minHours": "24",
                        "maxHours": "32",
                        "startDate": "2025-01-20",
                        "endDate": "2025-01-24",
                        "status": "Not Started"
                    },
                    {
                        "taskName": "UI/UX Design",
                        "resource": "UI Designer",
                        "minHours": "40",
                        "maxHours": "56",
                        "startDate": "2025-01-25",
                        "endDate": "2025-02-05",
                        "status": "Not Started"
                    }
                ],
                "subtotal": {
                    "minHours": 64,
                    "maxHours": 88
                }
            },
            {
                "name": "DEVELOPMENT PHASE",
                "tasks": [
                    {
                        "taskName": "Core Module Development",
                        "resource": "Senior Developer",
                        "minHours": "80",
                        "maxHours": "100",
                        "startDate": "2025-02-06",
                        "endDate": "2025-02-20",
                        "status": "Not Started"
                    }
                ],
                "subtotal": {
                    "minHours": 80,
                    "maxHours": 100
                }
            }
        ],
        "grandTotal": "GRAND TOTAL",
        "summary": {
            "totalMinHours": 144,
            "totalMaxHours": 188,
            "estimatedDurationWeeks": 4,
            "recommendedTeamSize": 3,
            "totalTasks": 3,
            "totalPhases": 2
        },
        "risks": [
            {
                "risk": "Technical complexity may exceed initial estimates",
                "phase": "Development Phase",
                "impact": "High",
                "mitigation": "Add buffer time and conduct thorough technical review"
            }
        ],
        "recommendations": [
            "Start with a detailed technical specification",
            "Consider using proven frameworks to reduce development time"
        ]
    }
    
    return json_to_excel(sample_data, "test_estimation.xlsx")

