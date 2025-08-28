import json
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import os

def load_json_data(json_file_path):
    """Load JSON data from file"""
    try:
        with open(json_file_path, 'r', encoding='utf-8') as file:
            data = json.load(file)
        print(f"✓ Successfully loaded JSON data from {json_file_path}")
        return data
    except FileNotFoundError:
        print(f"✗ Error: File {json_file_path} not found")
        return None
    except json.JSONDecodeError:
        print("✗ Error: Invalid JSON format")
        return None

def create_project_overview_sheet(wb, data):
    """Create Project Overview worksheet"""
    ws = wb.create_sheet("Project Overview")
    
    # Header styling
    header_font = Font(bold=True, size=12)
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    overview = data['project_overview']
    
    # Add project overview data
    overview_data = [
        ['Field', 'Value'],
        ['Project Title', overview['title']],
        ['Description', overview['description']],
        ['Total Min Hours', overview['total_min_hours']],
        ['Total Max Hours', overview['total_max_hours']],
        ['Estimated Duration (Weeks)', overview['estimated_duration_weeks']],
        ['Recommended Team Size', overview['team_size_recommended']]
    ]
    
    for row_num, row_data in enumerate(overview_data, 1):
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            if row_num == 1:  # Header row
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = header_fill
    
    # Auto-adjust column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 80
    
    print("✓ Created Project Overview sheet")

def create_tasks_sheet(wb, data):
    """Create Tasks worksheet"""
    ws = wb.create_sheet("Tasks Summary")
    
    # Header styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    # Headers
    headers = ['Task ID', 'Task Name', 'Description', 'Min Hours', 'Max Hours', 'Critical Path']
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
    
    # Add task data
    for row_num, task in enumerate(data['tasks'], 2):
        ws.cell(row=row_num, column=1, value=task['task_id'])
        ws.cell(row=row_num, column=2, value=task['task_name'])
        ws.cell(row=row_num, column=3, value=task['description'])
        ws.cell(row=row_num, column=4, value=task['task_total_min_hours'])
        ws.cell(row=row_num, column=5, value=task['task_total_max_hours'])
        ws.cell(row=row_num, column=6, value="Yes" if task['critical_path'] else "No")
    
    # Auto-adjust column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 60
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 15
    
    print("✓ Created Tasks Summary sheet")

def create_subtasks_sheet(wb, data):
    """Create Subtasks worksheet"""
    ws = wb.create_sheet("Subtasks Detail")
    
    # Header styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    # Headers
    headers = ['Task ID', 'Subtask ID', 'Subtask Name', 'Description', 'Min Hours', 'Max Hours', 'Skills Required', 'Complexity', 'Dependencies']
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
    
    # Add subtask data
    row_num = 2
    for task in data['tasks']:
        for subtask in task['subtasks']:
            ws.cell(row=row_num, column=1, value=task['task_id'])
            ws.cell(row=row_num, column=2, value=subtask['subtask_id'])
            ws.cell(row=row_num, column=3, value=subtask['subtask_name'])
            ws.cell(row=row_num, column=4, value=subtask['description'])
            ws.cell(row=row_num, column=5, value=subtask['min_hours'])
            ws.cell(row=row_num, column=6, value=subtask['max_hours'])
            ws.cell(row=row_num, column=7, value=', '.join(subtask['skills_required']))
            ws.cell(row=row_num, column=8, value=subtask['complexity'])
            ws.cell(row=row_num, column=9, value=', '.join(subtask['dependencies']) if subtask['dependencies'] else 'None')
            row_num += 1
    
    # Auto-adjust column widths
    column_widths = [12, 15, 35, 60, 12, 12, 30, 12, 25]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + i)].width = width
    
    print("✓ Created Subtasks Detail sheet")

def create_risks_sheet(wb, data):
    """Create Risks worksheet"""
    ws = wb.create_sheet("Risks & Considerations")
    
    # Header styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    # Headers
    headers = ['Risk', 'Impact', 'Mitigation']
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
    
    # Add risks data
    for row_num, risk in enumerate(data['risks_and_considerations'], 2):
        ws.cell(row=row_num, column=1, value=risk['risk'])
        ws.cell(row=row_num, column=2, value=risk['impact'])
        ws.cell(row=row_num, column=3, value=risk['mitigation'])
    
    # Auto-adjust column widths
    ws.column_dimensions['A'].width = 50
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 60
    
    print("✓ Created Risks & Considerations sheet")

def create_resources_sheet(wb, data):
    """Create Resources worksheet"""
    ws = wb.create_sheet("Resource Requirements")
    
    # Header styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    # Skills Summary section
    ws.cell(row=1, column=1, value="Skills Summary").font = Font(bold=True, size=14)
    
    # Skills headers
    skills_headers = ['Skill', 'Hours Required']
    for col_num, header in enumerate(skills_headers, 1):
        cell = ws.cell(row=3, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
    
    # Add skills data
    row_num = 4
    for skill, hours in data['resource_requirements']['skills_summary'].items():
        ws.cell(row=row_num, column=1, value=skill)
        ws.cell(row=row_num, column=2, value=hours)
        row_num += 1
    
    # Team Composition section
    start_row = row_num + 2
    ws.cell(row=start_row, column=1, value="Team Composition").font = Font(bold=True, size=14)
    
    # Team headers
    team_headers = ['Role', 'Skills', 'Estimated Hours']
    for col_num, header in enumerate(team_headers, 1):
        cell = ws.cell(row=start_row + 2, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
    
    # Add team data
    row_num = start_row + 3
    for member in data['resource_requirements']['team_composition']:
        ws.cell(row=row_num, column=1, value=member['role'])
        ws.cell(row=row_num, column=2, value=', '.join(member['skills']))
        ws.cell(row=row_num, column=3, value=member['estimated_hours'])
        row_num += 1
    
    # Auto-adjust column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 18
    
    print("✓ Created Resource Requirements sheet")

def create_assumptions_recommendations_sheet(wb, data):
    """Create Assumptions & Recommendations worksheet"""
    ws = wb.create_sheet("Assumptions & Recommendations")
    
    # Header styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    # Assumptions section
    ws.cell(row=1, column=1, value="Assumptions").font = Font(bold=True, size=14)
    
    for i, assumption in enumerate(data['assumptions'], 2):
        ws.cell(row=i, column=1, value=f"{i-1}. {assumption}")
    
    # Recommendations section
    start_row = len(data['assumptions']) + 4
    ws.cell(row=start_row, column=1, value="Recommendations").font = Font(bold=True, size=14)
    
    for i, recommendation in enumerate(data['recommendations'], start_row + 1):
        ws.cell(row=i, column=1, value=f"{i-start_row}. {recommendation}")
    
    # Auto-adjust column width
    ws.column_dimensions['A'].width = 80
    
    print("✓ Created Assumptions & Recommendations sheet")

def json_to_excel(json_data, output_file_path):
    """Main function to convert JSON to Excel"""
    print("🚀 Starting JSON to Excel conversion...")
    
    # Step 1: Load JSON data
    # data = load_json_data(json_file_path)
    if not json_data:
        return False
    
    # Step 2: Create workbook
    wb = Workbook()
    
    # Remove default sheet
    wb.remove(wb.active)
    
    # Step 3: Create all worksheets
    create_project_overview_sheet(wb, json_data)
    create_tasks_sheet(wb, json_data)
    create_subtasks_sheet(wb, json_data)
    create_risks_sheet(wb, json_data)
    create_resources_sheet(wb, json_data)
    create_assumptions_recommendations_sheet(wb, json_data)
    
    # Step 4: Save the workbook
    try:
        wb.save(output_file_path)
        print(f"✅ Successfully created Excel file: {output_file_path}")
        return True
    except Exception as e:
        print(f"✗ Error saving file: {e}")
        return False
